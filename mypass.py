from tkinter import *
from tkinter import messagebox
import tkinter.filedialog as fd
import os
import base64
import stat
import string
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
import logging
import ctypes


# UI Setup
bgColor = "#002120"
wdgtXPad = 5
wdgtYPad = 7



# Key
symkey = None



# Hard-coded items
folderName = "MyPass"
encryptedPassFilename = ".mypass.cred.xlsx"
tmpDecryptedPassFilename = "mypass.tmp.cred.xlsx"
saltedHashPassFilename = ".auth.xlsx"
logFilename = "mypass.log"
homePath = os.path.expanduser("~")
folderPath = os.path.join(os.path.expanduser("~"), folderName)
encryptedPassFilePath = os.path.join(homePath, folderName, encryptedPassFilename)
tmpDecryptedPassFilePath = os.path.join(homePath, folderName, tmpDecryptedPassFilename)
saltedHashPassFilePath = os.path.join(homePath, folderName, saltedHashPassFilename)
logPath = os.path.join(homePath, folderName, logFilename)
ctypes.windll.shcore.SetProcessDpiAwareness(1)


# Key check
passIsSame = False

# Import filename list
importedFiles = None


def initialize():
    """
    For first time creation:
    Application will create a folder created and called ~/.mypass.
    The file the passwords will be stored will be called .mypass.cred.xlsx.
    This .mypass.cred.xlsx will be an encrypted excel sheet with the Fernet library and will perform python functions to
    store into local variable after decrypting the sheet. It will only be decrypted during use of this application, after
    the password has been applied to it.

    The file to store the Master password is stored in .auth.mpdb.xlsx.
    The user's Master password is salted and hashed will be stored in a different file for future comparisons.

    After first time creation:
    Window will prompt for the user's master password and proceed to decrypt the file. For any file modifications, it must be done
    during the time the Application has its Main Window open.

    :return:
    """


    if not os.path.exists(folderPath):
        logging.info("Creating folder at " + folderPath)
        os.mkdir(folderPath)
    if not os.path.exists(logPath):
        fp = open(logPath, 'x')
        fp.close()
    logging.basicConfig(format='%(asctime)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p', filename=logPath,
                        encoding='utf-8', level=logging.DEBUG)
    logging.info("Initializing MyPass application.")
    if (not os.path.exists(encryptedPassFilePath) and not os.path.exists(saltedHashPassFilePath)):
        logging.info("Changing permissions of the folder at " + folderPath)
        os.chmod(folderPath, stat.S_IRWXU)
        logging.info("Setting master password at " + saltedHashPassFilePath)
        setPassword()
        logging.info("Starting up main window of MyPass to store credentials at " + encryptedPassFilePath)
        drawWindow()
    else:
        logging.info("Checking the master password is same to the one on file.")
        checkPassword()
        if passIsSame:
            logging.info("Decrypting the MyPass credentials located at " + encryptedPassFilePath + " with provided master password.")
            decryptFile()
            logging.info("Starting up main window of MyPass to store credentials at " + tmpDecryptedPassFilePath)
            drawWindow()
        else:
            logging.info("Exiting application.")
            sys.exit()


def drawWindow():
    """
    Draws the entire main window of MyPass GUI to add additional passwords
    :return:
    """
    window = Tk()

    # Menu Bar
    menubar = Menu(window)
    window.option_add('*tearOff', FALSE)
    fileMenu = Menu(menubar)
    fileMenu.add_command(label="Import Files", command=importFile)
    fileMenu.add_command(label="Open Vault", command=openVault)
    fileMenu.add_separator()
    fileMenu.add_command(label="Exit", underline=1, command=closeFunction)
    menubar.add_cascade(label="File", menu=fileMenu)
    window.config(menu=menubar)

    window.protocol("WM_DELETE_WINDOW", closeFunction) # Will execute this function before closing
    window.resizable(False, False)
    window.title("MyPass")
    window.configure(bg=bgColor)
    window.config(padx=50, pady=50)
    window.minsize(width=550, height=550)

    window.iconbitmap("resources/my-pass-logo.ico")
    canvas = Canvas(height=400, width=400)
    logoImg = PhotoImage(file="resources/my-pass-logo.png")
    canvas.create_image(200, 200, image=logoImg)
    canvas.configure(bg=bgColor)
    canvas.grid(column=1, row=1, columnspan=1)

    # Labels in columns 0
    websiteLabel = Label(text="Website:", font=("Arial", 14, "bold"), fg="white", bg=bgColor)
    websiteLabel.grid(column=0, row=2)
    emailLabel = Label(text="Email/Username:", font=("Arial", 14, "bold"), fg="white", bg=bgColor)
    emailLabel.grid(column=0, row=3)
    passwdLabel = Label(text="Password:", font=("Arial", 14, "bold"), fg="white", bg=bgColor)
    passwdLabel.grid(column=0, row=4)

    # Input fields in columns 1
    website = ""
    email = ""
    passwd = ""
    websiteInput = Entry(textvariable=website, width=80)
    websiteInput.grid(column=1, row=2, columnspan=2, padx=wdgtXPad, pady=wdgtYPad)
    websiteInput.focus()

    emailInput = Entry(textvariable=email, width=80)
    emailInput.grid(column=1, row=3, columnspan=2, padx=wdgtXPad, pady=wdgtYPad)

    passwdInput = Entry(textvariable=passwd, width=61)
    passwdInput.grid(column=1, row=4, padx=wdgtXPad, pady=wdgtYPad)

    addBtn = Button(text="ADD", width=60, bg="#1dd8f4", font=("Arial", 10, "bold"), command=lambda: addBtnCallback(websiteInput, emailInput, passwdInput))
    addBtn.grid(column=1, row=5, columnspan=2, padx=wdgtXPad, pady=wdgtYPad)

    # Button field in column 2
    genPass = Button(text="Generate Pass", font=("Arial", 10, "bold"), bg="#81f991", command=lambda: genPassword(passwdInput))
    genPass.grid(column=2, row=4, padx=wdgtXPad, pady=wdgtYPad)


    window.mainloop()




def genPassword(passwdField):
    """
    Randomly generates a password with cryptographic randomness
    :return:
    """
    thePass = ""
    totalAllowedCharacters = ['!','#','$','%', '&','(',')','*','+','_','^']
    totalAllowedCharacters += string.ascii_letters
    totalAllowedCharacters += string.digits
    totalCharacters = len(totalAllowedCharacters) - 1
    for x in range(25):
        singleChar = totalAllowedCharacters[int.from_bytes(os.urandom(32), "little") % totalCharacters]
        thePass += singleChar
    passwdField.delete(0,len(passwdField.get()))
    passwdField.insert(0,thePass)



def closeFunction():
    """
    Encrypts the file before exiting the application.
    :return:
    """
    logging.info("Encrypting user's credentials.")
    encryptFile()
    logging.info("Closing the MyPass application.")
    sys.exit()



def encryptFile():
    """
    Encrypts the entire temporary password file and writes to the encrypted pass file.
    :param userInput:
    :return:
    """
    if symkey:
        fernetKey = Fernet(symkey)
        with open(tmpDecryptedPassFilePath,'rb') as fileD:
            content = fileD.read()
        token = fernetKey.encrypt(content)
        with open(encryptedPassFilePath,'wb') as fileE:
            fileE.write(token)
        if os.path.exists(tmpDecryptedPassFilePath):
            os.remove(tmpDecryptedPassFilePath)
    else:
        logging.error("The symmetric key has failed to initialize during encryption!")




def decryptFile():
    """
    Decrypts the entire password file and stores into a temporary file.
    :param userInput: String input that the user has entered
    :return:
    """
    if symkey:
        tmpPassWb = Workbook()
        passWs = tmpPassWb.active
        passWs.title = "Passwords"
        tmpPassWb.save(tmpDecryptedPassFilePath)
        fernetKey = Fernet(symkey)
        with open(encryptedPassFilePath,'rb') as fileE:
            content = fileE.read()
        token = fernetKey.decrypt(content)
        with open(tmpDecryptedPassFilePath,'wb') as fileD:
            fileD.write(token)
    else:
        logging.error("The symmetric key has failed to initialize during decryption!")



def addBtnCallback(wEnt, eEnt, pEnt):
    """
    Will add all the information from the input fields and place it into the excel file.
    :return: None
    """
    passWb = load_workbook(tmpDecryptedPassFilePath)
    passWs = passWb["Passwords"]
    lastRow = passWs.max_row
    passWs.cell(column=1, row=lastRow+1, value=wEnt.get())
    passWs.cell(column=2, row=lastRow+1, value=eEnt.get())
    passWs.cell(column=3, row=lastRow+1, value=pEnt.get())
    passWb.save(tmpDecryptedPassFilePath)
    messagebox.showinfo(message="Credentials were added!")
    wEnt.delete(0, len(wEnt.get()))
    eEnt.delete(0, len(eEnt.get()))
    pEnt.delete(0, len(pEnt.get()))



def setPassword():
    """
    Opens a new window that inquires for user's new master password. This will include the salt as well.
    :return:
    """
    window = Tk()
    window.resizable(False, False)
    window.title("MyPass New Password")
    width=50
    window.configure(bg=bgColor)
    window.config(padx=50, pady=50)
    window.minsize(width=550, height=150)
    window.protocol("WM_DELETE_WINDOW", setPassCloseFunction)  # Will execute this function before closing

    # Labels in columns 0
    passwordLabel = Label(text="New Password:", font=("Arial", 14, "bold"), fg="white", bg=bgColor)
    passwordLabel.grid(column=0, row=0)
    confirmPassLabel = Label(text="Confirm Password:", font=("Arial", 14, "bold"), fg="white", bg=bgColor)
    confirmPassLabel.grid(column=0, row=1)

    password = ""
    confirmPass = ""
    passwordInput = Entry(textvariable=password, width=width, show='*')
    passwordInput.grid(column=1, row=0, columnspan=2, padx=wdgtXPad, pady=wdgtYPad)
    passwordInput.focus()

    confirmPassInput = Entry(textvariable=confirmPass, width=width, show='*')
    confirmPassInput.grid(column=1, row=1, columnspan=2, padx=wdgtXPad, pady=wdgtYPad)

    # Set password Button
    setPassBtn = Button(text="Set Password", font=("Arial", 14, "bold"), bg="#81f991", command=lambda: savePassword(window, passwordInput.get(), confirmPassInput.get()))
    setPassBtn.grid(column=2, row=2,  columnspan=2, padx=wdgtXPad, pady=wdgtYPad)
    window.mainloop()


def setPassCloseFunction():
    sys.exit()




def savePassword(window, passwd, confirmPasswd):
    """
    Save password into the .auth.mpdb.xlsx file.
    :param window: The New Password window
    :param passPath: Path of the excel sheet that stores the hashed password with the salt
    :param passwd: Users password
    :param confirmPasswd: Users password to make sure the user typed the password correctly
    :return:
    """
    global symkey
    if (passwd != confirmPasswd):
        logging.warning("Passwords did not match!")
    else:
        salt1 = os.urandom(32)
        passwdB = bytes(confirmPasswd, 'UTF-8')
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=salt1,
            iterations=390000,
        )
        symkey = base64.urlsafe_b64encode(kdf.derive(passwdB))
        safeSalt = base64.urlsafe_b64encode(salt1)
        hashesWb = Workbook()
        ws = hashesWb.active
        ws.title = "Master Pass"
        ws.cell(column=1,row=1,value=symkey)
        ws.cell(column=2, row=1, value=safeSalt)
        hashesWb.save(saltedHashPassFilePath)
        passwordWb = Workbook()
        passWs = passwordWb.active
        passWs.title = "Passwords"
        passwordWb.save(encryptedPassFilePath)
        window.destroy()
        tmpPassWb = Workbook()
        passWs = tmpPassWb.active
        passWs.title = "Passwords"
        tmpPassWb.save(tmpDecryptedPassFilePath)



def checkPassword():
    """
    Window will prompt user to enter the decryption password to access the file which is located in ~/MyPass/mypass.cred.xlsx.
    This file will be checked against ~/MyPass/mypass.cred.xlsx hash. If the file is tampered with, the passwords will no longer be accessible.
    :return: string input of the password in sha256 hash form
    """
    window = Tk()
    window.resizable(False, False)
    window.title("MyPass Password")
    width=50
    window.configure(bg=bgColor)
    window.config(padx=50, pady=50)
    window.minsize(width=550, height=150)

    # Labels in columns 0
    passwordLabel = Label(text="Password:", font=("Arial", 14, "bold"), fg="white", bg=bgColor)
    passwordLabel.grid(column=0, row=0)
    password = ""

    passwordInput = Entry(textvariable=password, width=width, show='*')
    passwordInput.grid(column=1, row=0, columnspan=2, padx=wdgtXPad, pady=wdgtYPad)
    passwordInput.focus()


    # Set password Button
    setPassBtn = Button(text="Enter", font=("Arial", 14, "bold"), bg="#81f991",command=lambda: compareMasterPassword(window, passwordInput.get()))
    setPassBtn.grid(column=2, row=2,  columnspan=2, padx=wdgtXPad, pady=wdgtYPad)
    window.mainloop()



def compareMasterPassword(window, password):
    """
    Compares the inputted password to the one inside the excel file and determines if it is the same.
    :param window: the main TK window that's open
    :param passPath: Path to the excel file
    :param password: the inputted password from the user
    :return: None
    """
    global passIsSame
    global symkey
    hashesWb = load_workbook(saltedHashPassFilePath)
    ws = hashesWb["Master Pass"]
    origPass = bytes(ws['A1'].value,'utf-8')
    salt = base64.urlsafe_b64decode(ws['B1'].value)
    passwdB = bytes(password, 'UTF-8')
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=390000,
    )
    enteredPass = base64.urlsafe_b64encode(kdf.derive(passwdB))
    if origPass == enteredPass:
        passIsSame = True
        symkey = enteredPass
        window.destroy()
    else:
        passIsSame = False
        window.destroy()



def importFile():
    """
    Callback function that imports the credentials that the user created in an excel sheet and stores into
    the temporary database for later encryption.
    :return:
    """
    importedFiles = fd.askopenfilenames(title="Select password files", initialdir=homePath,
                                        filetypes=[("Excel files", ".xlsx .xls")], multiple=True)
    if importedFiles:
        passWb = load_workbook(tmpDecryptedPassFilePath)
        passWs = passWb["Passwords"]
        lastRow = passWs.max_row
        for files in list(importedFiles):
            tempWb = load_workbook(files)
            tempWs = tempWb.active
            tempLastRow = tempWs.max_row
            for row in tempWs.iter_rows(min_row=1,max_col=3,max_row=tempLastRow):
                passWs.cell(column=1, row=lastRow + 1, value=row[0].value)
                passWs.cell(column=2, row=lastRow + 1, value=row[1].value)
                passWs.cell(column=3, row=lastRow + 1, value=row[2].value)
                lastRow += 1
        passWb.save(tmpDecryptedPassFilePath)
        messagebox.showinfo(message="Credentials were transferred!")
        importedFiles = None





def openVault():
    os.startfile(tmpDecryptedPassFilePath)


def main():
    initialize()


if __name__ == "__main__":
    main()

